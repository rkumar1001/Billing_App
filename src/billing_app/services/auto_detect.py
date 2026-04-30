"""Auto-detect field locations in an Excel breakdown + Word invoice pair.

We never want the end user to manually map cells. Instead we infer positions by
looking for visual cues; the detector is deliberately layout-agnostic so it can
cope with templates that differ in font, content, table shape, or even keep
their content inside text frames (Word drawings) rather than body tables.

Word strategy (works for any layout):
  - Walk ALL <w:p> paragraphs under the body, including those nested inside
    tables, drawings, and text frames. Compute each paragraph's full joined
    text (regardless of how Word split it across runs).
  - Match flexible patterns for: invoice_number, invoice_date, billing_period
    (date range), total_hours (numeric value near a "Total Hours" label or
    the right side of the billing-period row), and grand_total (currency with
    "Grand"/"Total"/"Balance Due" label, or last currency in doc as fallback).
  - Record the paragraph's doc-order index PLUS the matched substring so the
    editor can re-find and replace the text on every subsequent run.

Excel strategy is unchanged for .xlsx; .xls files (legacy BIFF) are read via
xlrd and mapped into the same ExcelDetection shape.
"""
from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator

from docx import Document
from docx.oxml.ns import qn
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .calendar_util import MONTH_NAMES


@dataclass
class ExcelDetection:
    sheet: str = ""
    month_number_cell: str | None = None
    invoice_date_cell: str | None = None
    dates_column_start: str | None = None
    dates_row_end: int | None = None  # last row that held a source-month date
    detected_month: int | None = None
    detected_year: int | None = None
    hourly_rate: float | None = None
    hours_per_day: float | None = None
    guards: int | None = None
    legacy_xls: bool = False  # True when the source was a .xls (BIFF) file
    formula_dates: bool = False  # True when col B has =TEXT(DATE(...)) style formulas
    text_dates: bool = False     # True when dates are stored as plain text strings
    text_date_format: str = ""   # strftime template derived from the source text
    warnings: list[str] = field(default_factory=list)

    def missing_fields(self) -> list[str]:
        out = []
        if self.month_number_cell is None:
            out.append("month_number_cell")
        if self.invoice_date_cell is None:
            out.append("invoice_date_cell")
        if self.dates_column_start is None:
            out.append("dates_column_start")
        return out

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "month_number_cell": self.month_number_cell,
            "invoice_date_cell": self.invoice_date_cell,
            "dates_column_start": self.dates_column_start,
            "dates_row_end": self.dates_row_end,
            "legacy_xls": self.legacy_xls,
            "formula_dates": self.formula_dates,
            "text_dates": self.text_dates,
            "text_date_format": self.text_date_format,
        }

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "ExcelDetection":
        return cls(
            sheet=data.get("sheet", ""),
            month_number_cell=data.get("month_number_cell"),
            invoice_date_cell=data.get("invoice_date_cell"),
            dates_column_start=data.get("dates_column_start"),
            dates_row_end=data.get("dates_row_end"),
            legacy_xls=bool(data.get("legacy_xls", False)),
            formula_dates=bool(data.get("formula_dates", False)),
            text_dates=bool(data.get("text_dates", False)),
            text_date_format=str(data.get("text_date_format", "")),
        )


@dataclass
class WordLoc:
    """A reference to a matched value inside a Word document.

    The primary lookup key is the text-based `match_text` — editing simply
    finds whichever paragraph contains that substring and replaces it. The
    `paragraph_order` is a tie-breaker when the same substring appears in
    multiple paragraphs (e.g. duplicated layouts for printing).

    Legacy fields (`table_index`, `row`, `col`) are preserved so that old
    `.billingapp.json` caches continue to load; they are no longer used by
    the new detection/edit flow.
    """
    # New primary fields
    paragraph_order: int = 0
    old_text: str = ""       # full joined text of the paragraph at detect time
    match_text: str = ""     # the specific substring that represents this field
    # Legacy fields (for cache back-compat only)
    table_index: int = -1
    row: int = 0
    col: int = 0
    paragraph_index: int = 0

    def to_dict(self) -> dict[str, Any]:
        return {
            "paragraph_order": self.paragraph_order,
            "old_text": self.old_text,
            "match_text": self.match_text,
            # Keep legacy keys in case an older app reads the file.
            "table": self.table_index,
            "row": self.row,
            "col": self.col,
            "paragraph": self.paragraph_index,
        }

    @classmethod
    def from_dict(cls, d: dict[str, Any]) -> "WordLoc":
        return cls(
            paragraph_order=int(d.get("paragraph_order", 0)),
            old_text=str(d.get("old_text", "")),
            match_text=str(d.get("match_text", "")),
            table_index=int(d.get("table", -1)),
            row=int(d.get("row", 0)),
            col=int(d.get("col", 0)),
            paragraph_index=int(d.get("paragraph", 0)),
        )


@dataclass
class WordDetection:
    invoice_number: WordLoc | None = None
    invoice_date: WordLoc | None = None
    billing_period: WordLoc | None = None  # Bella-style: single "X – Y" range
    from_date: WordLoc | None = None       # Hartford-style: separate "From Date" cell
    to_date: WordLoc | None = None         # Hartford-style: separate "To Date" cell
    total_hours: WordLoc | None = None
    grand_total: WordLoc | None = None
    warnings: list[str] = field(default_factory=list)

    _FIELDS = (
        "invoice_number", "invoice_date", "billing_period",
        "from_date", "to_date", "total_hours", "grand_total",
    )

    def missing_fields(self) -> list[str]:
        out: list[str] = []
        # An invoice has EITHER a billing_period (single range) OR a
        # from/to pair — only flag whichever style is missing entirely.
        for name in ("invoice_number", "invoice_date"):
            if getattr(self, name) is None:
                out.append(name)
        if self.billing_period is None and (self.from_date is None or self.to_date is None):
            out.append("billing_period (or from/to dates)")
        # total_hours / grand_total only matter when billing_period style
        # is in use; multi-row Hartford-style invoices keep these untouched.
        if self.billing_period is not None:
            for name in ("total_hours", "grand_total"):
                if getattr(self, name) is None:
                    out.append(name)
        return out

    def to_dict(self) -> dict[str, Any]:
        return {
            name: (getattr(self, name).to_dict() if getattr(self, name) else None)
            for name in self._FIELDS
        }

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "WordDetection":
        kwargs: dict[str, WordLoc | None] = {}
        for name in cls._FIELDS:
            v = data.get(name)
            kwargs[name] = WordLoc.from_dict(v) if v else None
        return cls(**kwargs)


# ---------------------------------------------------------------- Excel ----

def detect_excel(excel_path: Path) -> ExcelDetection:
    """Detect field locations in an Excel workbook.

    Accepts both `.xlsx`/`.xlsm` and legacy `.xls`. For .xls input we run
    a real-fidelity conversion (MS Excel COM on Windows, LibreOffice
    elsewhere) into a temp .xlsx, then run normal detection on it.
    """
    if excel_path.suffix.lower() == ".xls":
        from .xls_to_xlsx import XlsConvertError, convert_xls_to_xlsx
        # Place the temp .xlsx beside the source so sandboxed LibreOffice
        # snaps (which can't write to /tmp) still succeed.
        tmp_path = excel_path.parent / f".{excel_path.stem}.preview.xlsx"
        try:
            convert_xls_to_xlsx(excel_path, tmp_path)
            result = _detect_xlsx(tmp_path)
        except XlsConvertError as e:
            result = ExcelDetection(legacy_xls=True)
            result.warnings.append(str(e))
            return result
        finally:
            try:
                if tmp_path.exists():
                    tmp_path.unlink()
            except OSError:
                pass
        result.legacy_xls = True
        return result
    return _detect_xlsx(excel_path)


def _detect_xlsx(xlsx_path: Path) -> ExcelDetection:
    from .file_role import pick_data_sheet
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    try:
        sheet_name = pick_data_sheet(wb)
        ws = wb[sheet_name]
        result = ExcelDetection(sheet=ws.title)

        # Month number: integer 1..12 in top 5 rows, rightmost wins.
        month_cell: str | None = None
        best_col = -1
        for row in range(1, 6):
            for col in range(1, 20):
                val = ws.cell(row=row, column=col).value
                if isinstance(val, int) and 1 <= val <= 12 and col > best_col:
                    # Skip the "BELLA VISTA COMMUNITY"-style text-number mix.
                    left = ws.cell(row=row, column=col - 1).value if col > 1 else None
                    if left in (None, ""):
                        best_col = col
                        month_cell = f"{get_column_letter(col)}{row}"
        result.month_number_cell = month_cell

        # Invoice date: a filled (non-default) cell containing a date, in top 5.
        date_cell: str | None = None
        fallback_date_cell: str | None = None
        for row in range(1, 6):
            for col in range(1, 20):
                cell = ws.cell(row=row, column=col)
                val = cell.value
                is_date = isinstance(val, (date, datetime))
                if not is_date:
                    continue
                fill = cell.fill
                coloured = (
                    fill
                    and fill.fgColor
                    and fill.fgColor.type == "rgb"
                    and fill.fgColor.rgb
                    and fill.fgColor.rgb not in ("00000000", "FFFFFFFF", "FFFFFF", "00FFFFFF")
                )
                addr = f"{get_column_letter(col)}{row}"
                if coloured and date_cell is None:
                    date_cell = addr
                if fallback_date_cell is None:
                    fallback_date_cell = addr
        result.invoice_date_cell = date_cell or fallback_date_cell

        # Dates column: look for a column where consecutive rows hold dates
        # of the same month, starting with day 1.
        best: tuple[str, int, int, int] | None = None  # (col_letter, start_row, month, year)
        for col in range(1, 20):
            run_start = None
            run_month = None
            run_year = None
            run_len = 0
            for row in range(1, 60):
                val = ws.cell(row=row, column=col).value
                d = _as_date(val)
                if d is None:
                    if run_len >= 20:  # good enough
                        break
                    run_start = None
                    run_month = None
                    run_len = 0
                    continue
                if run_start is None:
                    if d.day != 1:
                        continue
                    run_start = row
                    run_month = d.month
                    run_year = d.year
                    run_len = 1
                else:
                    expected_day = run_len + 1
                    if d.month == run_month and d.day == expected_day:
                        run_len += 1
                    else:
                        break
            if run_len >= 20:
                col_letter = get_column_letter(col)
                if best is None or run_len > best[2]:
                    best = (col_letter, run_start, run_len, run_year or 0)
                    result.dates_column_start = f"{col_letter}{run_start}"
                    result.dates_row_end = run_start + run_len - 1
                    result.detected_month = run_month
                    result.detected_year = run_year

        # If no literal-date column, check for formula-driven dates BEFORE
        # text-date detection — the Hartford-style template uses
        # =TEXT(DATE(...)) formulas whose cached display value is a string,
        # which would otherwise mis-trigger the text-date branch.
        if result.dates_column_start is None:
            formula_wb = load_workbook(filename=str(xlsx_path), data_only=False)
            try:
                ws_raw = formula_wb[ws.title] if ws.title in formula_wb.sheetnames else formula_wb.active
                first_formula_row: int | None = None
                last_formula_row: int | None = None
                for col in range(1, 12):
                    first_formula_row = None
                    last_formula_row = None
                    for row in range(1, 60):
                        v = ws_raw.cell(row=row, column=col).value
                        if isinstance(v, str) and v.startswith("=") and "DATE(" in v.upper():
                            if first_formula_row is None:
                                first_formula_row = row
                            last_formula_row = row
                    if first_formula_row is not None and last_formula_row - first_formula_row >= 5:
                        col_letter = get_column_letter(col)
                        result.dates_column_start = f"{col_letter}{first_formula_row}"
                        result.dates_row_end = last_formula_row
                        result.formula_dates = True
                        sample = ws_raw.cell(
                            row=first_formula_row, column=col,
                        ).value or ""
                        ym = re.search(r"DATE\(\s*\"?(\d{4})", sample)
                        if ym:
                            result.detected_year = int(ym.group(1))
                        break
            finally:
                formula_wb.close()

        # If neither literal nor formula dates, try a plain text-date column.
        # Templates that bake the date into a string (e.g. " Wednesday,
        # April-1 ,2026") fall here.
        if result.dates_column_start is None:
            text_date_re = re.compile(
                r"\s*(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\w*\s*,\s*"
                r"(?:January|February|March|April|May|June|July|August|"
                r"September|October|November|December|Jan|Feb|Mar|Apr|"
                r"Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*[-/ ]\s*"
                r"(?P<day>\d{1,2})\s*[, ]*\s*(?P<year>\d{4})",
                re.IGNORECASE,
            )
            for col in range(1, 12):
                run_start: int | None = None
                run_len = 0
                run_year = None
                run_month: int | None = None
                for row in range(1, 60):
                    v = ws.cell(row=row, column=col).value
                    if not isinstance(v, str):
                        if run_len >= 20:
                            break
                        run_start = None
                        run_len = 0
                        continue
                    m = text_date_re.match(v)
                    if not m:
                        if run_len >= 20:
                            break
                        run_start = None
                        run_len = 0
                        continue
                    day = int(m.group("day"))
                    if run_start is None:
                        if day != 1:
                            continue
                        run_start = row
                        run_len = 1
                        run_year = int(m.group("year"))
                    else:
                        if day == run_len + 1:
                            run_len += 1
                        else:
                            break
                if run_len >= 20:
                    col_letter = get_column_letter(col)
                    result.dates_column_start = f"{col_letter}{run_start}"
                    result.dates_row_end = run_start + run_len - 1
                    if result.detected_year is None:
                        result.detected_year = run_year
                    result.text_dates = True
                    sample = ws.cell(row=run_start, column=col).value or ""
                    fmt = _derive_text_date_format(sample)
                    if fmt:
                        result.text_date_format = fmt
                    # Try to derive the source month from the sample text.
                    if result.detected_month is None:
                        for idx, name in enumerate(MONTH_NAMES, start=1):
                            if re.search(rf"\b{name}\b", sample, re.IGNORECASE):
                                result.detected_month = idx
                                break
                    break

        # Read rate, hours per day, guards from the first date row.
        if result.dates_column_start:
            start_cell = result.dates_column_start
            col_letter = re.match(r"([A-Z]+)", start_cell).group(1)
            col_idx = _col_to_index(col_letter)
            start_row = int(start_cell[len(col_letter):])

            # Reload without data_only so we can tell formula cells apart
            # from literal ones. Formula cells are derived (e.g. regular
            # hours = hours * guards), so they're poor signals for rate
            # or hours-per-day detection.
            try:
                wb_raw = load_workbook(filename=str(xlsx_path), data_only=False)
                ws_raw = wb_raw[ws.title] if ws.title in wb_raw.sheetnames else wb_raw.active
            except Exception:  # noqa: BLE001
                ws_raw = None

            # Header-based mapping: prefer columns whose row-above header
            # explicitly names the field (works regardless of column order).
            label_map = {
                "hours_per_day": _re_header(r"hours?\s*per\s*day"),
                "guards": _re_header(r"(?:number\s*of\s*)?guards?"),
                "hourly_rate": _re_header(r"hourly\s*rate|rate"),
            }
            header_row = start_row - 1
            for col in range(col_idx + 1, col_idx + 11):
                header = ws.cell(row=header_row, column=col).value
                if not isinstance(header, str):
                    continue
                txt = header.strip().lower()
                for field_name, pat in label_map.items():
                    if pat.search(txt):
                        val = ws.cell(row=start_row, column=col).value
                        if isinstance(val, (int, float)) and val > 0:
                            cur = getattr(result, field_name)
                            if cur is None:
                                setattr(
                                    result, field_name,
                                    int(val) if field_name == "guards" else float(val),
                                )
                        break

            # Fallback positional sweep — but skip cells whose underlying
            # value is a formula (they're derived, not source data).
            for col in range(col_idx + 1, col_idx + 8):
                val = ws.cell(row=start_row, column=col).value
                raw = ws_raw.cell(row=start_row, column=col).value if ws_raw else None
                is_formula = isinstance(raw, str) and raw.startswith("=")
                if is_formula:
                    continue
                if isinstance(val, (int, float)) and val > 0:
                    if result.hours_per_day is None and 1 <= val <= 24 and isinstance(val, int):
                        result.hours_per_day = float(val)
                    elif result.guards is None and 1 <= val <= 50 and isinstance(val, int):
                        result.guards = int(val)
                    elif result.hourly_rate is None and val >= 5:
                        result.hourly_rate = float(val)
            if ws_raw is not None:
                wb_raw.close()

            if result.hours_per_day is None:
                result.hours_per_day = 0.0
            if result.guards is None:
                result.guards = 1
            if result.hourly_rate is None:
                result.hourly_rate = 0.0

        return result
    finally:
        wb.close()


def _re_header(pattern: str) -> "re.Pattern[str]":
    return re.compile(rf"\b{pattern}\b", re.IGNORECASE)


def _derive_text_date_format(sample: str) -> str | None:
    """Convert a sample date string like ' Wednesday, April-1 ,2026' into a
    strftime template that reproduces it for any other date.

    Returns None if the sample's structure can't be recognised.
    """
    import platform
    day_directive = "%#d" if platform.system() == "Windows" else "%-d"
    weekday_re = re.compile(
        r"(?P<wd>Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday|"
        r"Mon|Tue|Wed|Thu|Fri|Sat|Sun)",
        re.IGNORECASE,
    )
    month_re = re.compile(
        r"(?P<mo>January|February|March|April|May|June|July|August|September|"
        r"October|November|December|Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Oct|"
        r"Nov|Dec)",
        re.IGNORECASE,
    )
    wd_m = weekday_re.search(sample)
    mo_m = month_re.search(sample, pos=wd_m.end() if wd_m else 0)
    if not wd_m or not mo_m:
        return None
    # Day: 1-2 digits AFTER the month.
    day_m = re.search(r"\b(\d{1,2})\b", sample[mo_m.end():])
    # Year: 4-digit number AFTER the day.
    year_m = None
    if day_m:
        year_m = re.search(r"\b(\d{4})\b", sample[mo_m.end() + day_m.end():])
    if not day_m or not year_m:
        return None

    # Long vs abbreviated forms.
    wd_dir = "%A" if len(wd_m.group("wd")) > 3 else "%a"
    mo_dir = "%B" if len(mo_m.group("mo")) > 3 else "%b"

    # Stitch the original literal text back together with directives.
    out = sample[: wd_m.start()] + wd_dir
    out += sample[wd_m.end(): mo_m.start()] + mo_dir
    out += sample[mo_m.end(): mo_m.end() + day_m.start()] + day_directive
    abs_day_end = mo_m.end() + day_m.end()
    abs_year_start = mo_m.end() + day_m.end() + year_m.start()
    abs_year_end = mo_m.end() + day_m.end() + year_m.end()
    out += sample[abs_day_end: abs_year_start] + "%Y"
    out += sample[abs_year_end:]
    return out


def _as_date(val: Any) -> date | None:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _col_to_index(letters: str) -> int:
    acc = 0
    for c in letters.upper():
        acc = acc * 26 + (ord(c) - ord("A") + 1)
    return acc


# ---------------------------------------------------------------- Word -----

# Date range (e.g. 4/1/2026 – 4/30/2026, also accepts - — or "to"/"through").
_DATE_RANGE_RE = re.compile(
    r"\b\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}\b"
    r"\s*(?:[–\-—]|to|through)\s*"
    r"\b\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}\b",
    re.IGNORECASE,
)
# Currency: $X, X USD, Rs. X, ₹X with optional commas/decimals.
_CURRENCY_RE = re.compile(
    r"(?:\$|₹|Rs\.?\s?)\s*\d[\d,]*(?:\.\d+)?"
    r"|\b\d[\d,]*(?:\.\d+)?\s*(?:USD|INR)\b",
    re.IGNORECASE,
)
# Standalone date (numeric or long-form).
_DATE_ONLY_RE = re.compile(
    r"(?:\b\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}\b"
    r"|\b\d{4}-\d{2}-\d{2}\b"
    r"|(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|"
    r"Dec(?:ember)?)\s+\d{1,2},?\s*\d{4}"
    r"|\b\d{1,2}\s+(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|"
    r"Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|"
    r"Nov(?:ember)?|Dec(?:ember)?)\s+\d{4}\b)",
    re.IGNORECASE,
)
# Plain numeric cell contents.
_NUMERIC_RE = re.compile(r"^\s*\d+(?:\.\d+)?\s*$")
# Inline numeric (extract a lone number from a paragraph).
_INLINE_NUMBER_RE = re.compile(r"\b\d+(?:\.\d+)?\b")

# Label patterns we recognise as anchors for specific fields.
_TOTAL_HOURS_LABEL_RE = re.compile(
    r"\b(?:total\s*hours?|hours?\s*total|total\s*hrs?)\b", re.IGNORECASE
)
_GRAND_TOTAL_LABEL_RE = re.compile(
    r"\b(?:grand\s*total|amount\s*due|balance\s*due|total\s*due|total)\b",
    re.IGNORECASE,
)
_INVOICE_DATE_LABEL_RE = re.compile(
    r"\b(?:invoice\s*date|date\s*of\s*invoice|date)\b", re.IGNORECASE
)


def _iter_paragraphs(doc) -> Iterator[tuple[int, Any]]:
    """Yield (paragraph_order, w:p element) for every LEAF paragraph in the
    body, regardless of whether it lives in a table cell, drawing, or text
    frame. Paragraphs that contain other paragraphs (i.e. outer wrappers
    around text frames) are skipped so we don't get back mega-blobs of
    concatenated text alongside the tight inner matches.
    """
    p_tag = qn("w:p")
    for idx, p_elem in enumerate(doc.element.body.iter(p_tag)):
        # Skip container paragraphs: any <w:p> that has a descendant <w:p>
        # is an outer wrapper around a drawing/text frame.
        has_nested = False
        for child in p_elem.iter(p_tag):
            if child is not p_elem:
                has_nested = True
                break
        if has_nested:
            continue
        yield idx, p_elem


def _joined_text(p_elem) -> str:
    """Concatenate all <w:t> run texts inside a paragraph into one string."""
    parts = [t.text or "" for t in p_elem.iter(qn("w:t"))]
    return "".join(parts)


def detect_word(docx_path: Path, invoice_number_str: str) -> WordDetection:
    doc = Document(str(docx_path))
    result = WordDetection()

    # Snapshot: every paragraph's joined text, plus the preceding paragraphs
    # so we can look "above" a cell for label context.
    paragraphs: list[tuple[int, str]] = []
    for order, p_elem in _iter_paragraphs(doc):
        text = _joined_text(p_elem)
        stripped = text.strip()
        if stripped:
            paragraphs.append((order, text))

    def _set_loc(field_name: str, order: int, full: str, match: str) -> None:
        if getattr(result, field_name) is None:
            setattr(result, field_name, WordLoc(
                paragraph_order=order, old_text=full, match_text=match,
            ))

    # --- invoice_number (exact-ish match against the parsed filename number)
    if invoice_number_str:
        # Build a tolerant pattern: e.g. "HART23" -> "HART\s*23" so the doc's
        # "HART 23" matches, AND we record the EXACT substring found so the
        # editor replaces it precisely (preserving the space if present).
        m_split = re.match(r"^([A-Za-z]+)\s*(\d+)$", invoice_number_str)
        joined_target = invoice_number_str.replace(" ", "")
        for order, text in paragraphs:
            compact = text.replace(" ", "")
            if joined_target not in compact:
                continue
            real = None
            if m_split:
                real = re.search(
                    rf"{re.escape(m_split.group(1))}\s*{m_split.group(2)}",
                    text,
                )
            if real is None:
                real = re.search(re.escape(invoice_number_str), text)
            match = real.group(0) if real else invoice_number_str
            _set_loc("invoice_number", order, text, match)
            break

    # --- billing_period (date range, Bella-style)
    for order, text in paragraphs:
        m = _DATE_RANGE_RE.search(text)
        if m:
            _set_loc("billing_period", order, text, m.group(0))
            break

    # --- from_date / to_date (Hartford-style: two separate cells in a
    # multi-row line-items table). Only attempt this when no single-paragraph
    # billing_period was found, since the two styles are mutually exclusive.
    if result.billing_period is None:
        # Heuristic: walk paragraphs, tracking when we've passed a "From" /
        # "From Date" header. The first standalone-date paragraph after that
        # is from_date; the next one is to_date. A "standalone date" paragraph
        # is one whose joined text is JUST a single date (no range, no label).
        seen_from_label = False
        seen_to_label = False
        from_label_re = re.compile(r"\bfrom(?:\s+date)?\b", re.IGNORECASE)
        to_label_re = re.compile(r"\bto(?:\s+date)?\b", re.IGNORECASE)
        # We need to handle the case where "From" and "Date" are split into
        # two consecutive paragraphs (template shows "From / Date" stacked).
        prev_text = ""
        for order, text in paragraphs:
            stripped = text.strip()
            # Mark labels.
            combined_prev = (prev_text + " " + stripped).strip()
            if not seen_from_label and (
                from_label_re.search(stripped)
                or from_label_re.search(combined_prev)
            ):
                seen_from_label = True
            elif seen_from_label and not seen_to_label and (
                # 'To' alone, OR 'To Date'.
                re.fullmatch(r"\s*to\s*", stripped, re.IGNORECASE)
                or to_label_re.search(stripped)
            ):
                seen_to_label = True
            # Once we've passed the headers, watch for the FIRST two
            # standalone-date paragraphs.
            if seen_from_label:
                m = _DATE_ONLY_RE.fullmatch(stripped) or (
                    _DATE_ONLY_RE.match(stripped) if _DATE_ONLY_RE.match(stripped)
                    and _DATE_ONLY_RE.match(stripped).group(0) == stripped
                    else None
                )
                if m:
                    if result.from_date is None:
                        _set_loc("from_date", order, text, m.group(0))
                    elif result.to_date is None:
                        _set_loc("to_date", order, text, m.group(0))
                        break
            prev_text = stripped

    # --- invoice_date
    # Prefer: a paragraph whose nearby label says "Invoice Date" / "Date".
    #   Strategy: look for a date in a paragraph whose own text, or the
    #   immediately preceding non-empty paragraph, contains the label.
    prev_text = ""
    for order, text in paragraphs:
        if result.invoice_date is not None:
            break
        has_label_here = bool(_INVOICE_DATE_LABEL_RE.search(text))
        has_label_prev = bool(_INVOICE_DATE_LABEL_RE.search(prev_text))
        if has_label_here or has_label_prev:
            m = _DATE_ONLY_RE.search(text)
            # Skip if the only "date" here is the year inside a billing period.
            if m and not _DATE_RANGE_RE.search(text):
                _set_loc("invoice_date", order, text, m.group(0))
        prev_text = text
    # Fallback: first date-like paragraph that isn't a range.
    if result.invoice_date is None:
        for order, text in paragraphs:
            if _DATE_RANGE_RE.search(text):
                continue
            m = _DATE_ONLY_RE.search(text)
            if m:
                _set_loc("invoice_date", order, text, m.group(0))
                break

    # --- grand_total (prefer currency near a Grand Total / Amount Due label)
    # Scan twice: first pass picks currency in paragraphs whose text (or the
    # paragraph just before) contains the label; second pass falls back to the
    # LAST currency anywhere in the doc.
    prev_text = ""
    for order, text in paragraphs:
        if result.grand_total is not None:
            break
        has_label = bool(
            _GRAND_TOTAL_LABEL_RE.search(text)
            or _GRAND_TOTAL_LABEL_RE.search(prev_text)
        )
        if has_label:
            m = _CURRENCY_RE.search(text)
            if m:
                _set_loc("grand_total", order, text, m.group(0))
        prev_text = text
    if result.grand_total is None:
        last: tuple[int, str, str] | None = None
        for order, text in paragraphs:
            for m in _CURRENCY_RE.finditer(text):
                last = (order, text, m.group(0))
        if last:
            _set_loc("grand_total", *last)

    # --- total_hours
    # Preferred signal: a paragraph whose own text contains BOTH the label
    # and a number (e.g. "Total Hours: 330"). When the label lives in its
    # own paragraph (common in column-header layouts), scan forward from the
    # billing_period paragraph and pick the last bare number before we hit
    # the grand-total area.
    for order, text in paragraphs:
        if result.total_hours is not None:
            break
        if _TOTAL_HOURS_LABEL_RE.search(text):
            # Label is in this paragraph; is a number there too?
            masked = _TOTAL_HOURS_LABEL_RE.sub("", text)
            m = _INLINE_NUMBER_RE.search(masked)
            if m:
                _set_loc("total_hours", order, text, m.group(0))
    if result.total_hours is None and result.billing_period is not None:
        bp = result.billing_period
        collected: list[tuple[int, str, str]] = []
        # Scan the billing_period paragraph itself (minus the date range),
        # then any following paragraphs until we reach currency / grand-total.
        bp_text_masked = bp.old_text.replace(bp.match_text, " " * len(bp.match_text))
        for m in _INLINE_NUMBER_RE.finditer(bp_text_masked):
            collected.append((bp.paragraph_order, bp.old_text, m.group(0)))
        started = False
        for order, text in paragraphs:
            if order == bp.paragraph_order:
                started = True
                continue
            if not started:
                continue
            # Stop only once we reach an explicit grand-total area (label or
            # currency that looks like a grand total). Skipping intermediate
            # currency paragraphs means we still see "$20 hourly rate" and
            # then catch the bare "330" total-hours value that follows.
            if _GRAND_TOTAL_LABEL_RE.search(text):
                break
            if _CURRENCY_RE.search(text):
                # Hourly rate / daily amount cells — skip their digits.
                continue
            for m in _INLINE_NUMBER_RE.finditer(text):
                collected.append((order, text, m.group(0)))
        if collected:
            # Prefer the biggest value seen — total hours is typically larger
            # than neighbouring per-day counts.
            def _numval(tup: tuple[int, str, str]) -> float:
                try:
                    return float(tup[2])
                except ValueError:
                    return 0.0
            chosen = max(collected, key=_numval)
            _set_loc("total_hours", *chosen)

    return result


# ---------------------------------------------------------------- cache ----

CACHE_FILENAME = ".billingapp.json"


@dataclass
class DetectionCache:
    excel: ExcelDetection
    word: WordDetection

    def to_dict(self) -> dict[str, Any]:
        return {"excel": self.excel.to_dict(), "word": self.word.to_dict()}

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "DetectionCache":
        return cls(
            excel=ExcelDetection.from_dict(data.get("excel", {})),
            word=WordDetection.from_dict(data.get("word", {})),
        )


def load_cache(folder: Path) -> DetectionCache | None:
    path = folder / CACHE_FILENAME
    if not path.exists():
        return None
    try:
        return DetectionCache.from_dict(json.loads(path.read_text(encoding="utf-8")))
    except (json.JSONDecodeError, OSError, KeyError, ValueError):
        return None


def save_cache(folder: Path, cache: DetectionCache) -> None:
    path = folder / CACHE_FILENAME
    try:
        path.write_text(json.dumps(cache.to_dict(), indent=2), encoding="utf-8")
    except OSError:
        pass
