"""Classify workbooks by their content/purpose.

Different invoice clients use different auxiliary forms in addition to the
main breakdown spreadsheet. This module looks at the contents of an .xlsx
(or pre-converted .xls) workbook and decides which "role" it plays so the
generator can apply the right edit recipe to each file.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

Role = Literal["breakdown", "payment_request_form", "g703", "unknown"]

# Distinctive phrase fingerprints for each role. Keep them tight — false
# positives are worse than misses (a misclassified file is silently edited
# with the wrong recipe).
_BREAKDOWN_HINTS = (
    re.compile(r"\bbreakdown\s+of\s+services\b", re.IGNORECASE),
    re.compile(r"\bdaily\s+amount\b", re.IGNORECASE),
    re.compile(r"\bhours\s+per\s+day\b", re.IGNORECASE),
)
_PRF_HINTS = (
    re.compile(r"\bpayment\s+request\s+form\b", re.IGNORECASE),
    re.compile(r"\bperiod\s+ending\b", re.IGNORECASE),
    re.compile(r"\boriginal\s+contract\s+amount\b", re.IGNORECASE),
)
_G703_HINTS = (
    re.compile(r"\bcontinuation\s+sheet\b", re.IGNORECASE),
    re.compile(r"\bdocument\s+G703\b", re.IGNORECASE),
    re.compile(r"\bapplication\s+number\b", re.IGNORECASE),
    re.compile(r"\bdescription\s+of\s+work\b", re.IGNORECASE),
)


def classify_workbook(xlsx_path: Path) -> Role:
    """Return the role of `xlsx_path`. Always operates on .xlsx — caller is
    responsible for converting legacy .xls first.
    """
    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True)
    except Exception:  # noqa: BLE001
        return "unknown"

    text_chunks: list[str] = []
    try:
        for ws in wb.worksheets:
            # Sweep the first ~30 rows of every sheet — the role-defining
            # headers always live near the top of these forms.
            for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                for v in row:
                    if isinstance(v, str) and v.strip():
                        text_chunks.append(v)
    finally:
        wb.close()

    blob = " | ".join(text_chunks)

    breakdown_hits = sum(1 for r in _BREAKDOWN_HINTS if r.search(blob))
    prf_hits = sum(1 for r in _PRF_HINTS if r.search(blob))
    g703_hits = sum(1 for r in _G703_HINTS if r.search(blob))

    # G703 wins over PRF when both fire — G703 is more specific (it always
    # contains "DESCRIPTION OF WORK" + "APPLICATION NUMBER" in tandem).
    if g703_hits >= 2:
        return "g703"
    if prf_hits >= 2:
        return "payment_request_form"
    if breakdown_hits >= 2:
        return "breakdown"
    return "unknown"


def pick_data_sheet(wb) -> str:
    """Return the name of the workbook's most data-rich sheet.

    Default openpyxl behaviour is to use `wb.active`, which can land on an
    empty Sheet1 when the user saved the file with a different tab focused.
    Pick whichever non-trivial sheet has the most populated cells.
    """
    best_name = wb.active.title
    best_score = -1
    for ws in wb.worksheets:
        # Quick heuristic: count populated cells in the first 60 rows.
        score = 0
        for row in ws.iter_rows(min_row=1, max_row=60, values_only=True):
            for v in row:
                if v not in (None, ""):
                    score += 1
        if score > best_score:
            best_score = score
            best_name = ws.title
    return best_name
