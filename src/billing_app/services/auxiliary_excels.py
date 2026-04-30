"""Edit recipes for auxiliary spreadsheets that ride alongside the main
breakdown: AIA G703 continuation sheets and Fassberg-style Payment Request
Forms. Each recipe hunts for a small set of labelled cells by content
(case-insensitive label match) and bumps just the date- and
sequence-number-style fields each month, leaving every other value alone.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from .calendar_util import days_in_month
from .file_role import pick_data_sheet


def update_payment_request_form(
    path: Path,
    *,
    target_year: int,
    target_month: int,
    invoice_date: date,
) -> list[str]:
    """Bump the Period Ending + Date fields on a Fassberg-style PRF.

    Returns a list of any field names we couldn't find.
    """
    return _update_with_label_map(
        path,
        labels={
            "period ending": ("date", date(target_year, target_month, days_in_month(target_year, target_month))),
            "date": ("date", invoice_date),
            # Some templates use "Period to" / "Period From"
            "period to": ("date", date(target_year, target_month, days_in_month(target_year, target_month))),
        },
    )


def update_g703(
    path: Path,
    *,
    target_year: int,
    target_month: int,
    invoice_date: date,
    invoice_number_str: str = "",
) -> list[str]:
    """Bump the Application Number, Application Date, and Period To on an
    AIA G703 continuation sheet.

    The application-number sequence on the G703 is independent of the Word
    invoice number, so we read whatever is currently in the cell, bump
    the trailing integer by 1, and write the new value back. The
    `invoice_number_str` argument is used only as a last-resort fallback
    when the source cell is empty.
    """
    bumped = _read_and_bump_application_number(path, fallback=invoice_number_str)
    return _update_with_label_map(
        path,
        labels={
            "application number": ("invoice_number", bumped),
            "application date": ("date", invoice_date),
            "period to": ("date", date(target_year, target_month, days_in_month(target_year, target_month))),
        },
    )


def _read_and_bump_application_number(path: Path, *, fallback: str) -> str:
    """Open `path`, find the cell holding the Application Number value,
    parse a trailing integer (e.g. ``HART20``), and return the bumped
    form (``HART21``). If the cell is empty or unparseable, return
    `fallback`."""
    wb = load_workbook(filename=str(path), data_only=True)
    try:
        ws = wb[pick_data_sheet(wb)]
        label_re = re.compile(r"^\s*application\s*number\s*[:\-]?\s*$", re.IGNORECASE)
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30)):
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                if label_re.match(v.strip().rstrip(":")) or label_re.match(v):
                    target = _value_cell_after_label(
                        ws, cell.row, cell.column, kind="invoice_number",
                    )
                    if target is not None and isinstance(target.value, str):
                        m = re.match(
                            r"^\s*([A-Za-z]+)(\s*)(\d+)\s*$", target.value,
                        )
                        if m:
                            prefix, sp, num = m.group(1), m.group(2), int(m.group(3))
                            return f"{prefix}{sp}{num + 1}"
                    return fallback
        return fallback
    finally:
        wb.close()


def _update_with_label_map(
    path: Path,
    *,
    labels: dict[str, tuple[str, object]],
) -> list[str]:
    """Open `path`, scan the data-bearing sheet for the given labels, and
    write the corresponding values into the cell directly to the right
    (then the cell directly below, as a fallback) of each label.
    """
    wb = load_workbook(filename=str(path))
    unresolved: list[str] = []
    try:
        ws = wb[pick_data_sheet(wb)]
        # Normalise labels for matching.
        label_patterns = {
            re.compile(rf"^\s*{re.escape(label)}\s*[:\-]?\s*$", re.IGNORECASE): kind_value
            for label, kind_value in labels.items()
        }
        applied: set[str] = set()
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60)):
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                stripped = v.strip().rstrip(":").strip()
                for pattern, (kind, target_value) in label_patterns.items():
                    label_key = pattern.pattern
                    if label_key in applied:
                        continue
                    if pattern.match(v) or pattern.match(stripped + ":") or _starts_with_label(v, pattern):
                        target_cell = _value_cell_after_label(
                            ws, cell.row, cell.column, kind=kind,
                        )
                        if target_cell is None:
                            continue
                        _write_value(target_cell, kind, target_value)
                        applied.add(label_key)
                        break
        # Anything we didn't apply is unresolved.
        for pattern, (kind, _val) in label_patterns.items():
            if pattern.pattern not in applied:
                unresolved.append(pattern.pattern)
        wb.save(str(path))
    finally:
        wb.close()
    return unresolved


def _starts_with_label(text: str, pattern: re.Pattern) -> bool:
    """Match labels that include surrounding whitespace, dashes, or
    colon punctuation (e.g. '   Period Ending: ')."""
    cleaned = re.sub(r"[\s:\-]+$", "", text.strip())
    return bool(pattern.match(cleaned + ":"))


def _value_cell_after_label(ws, row: int, col: int, *, kind: str = ""):
    """Locate the cell that holds the *value* for a label.

    Forms typically place the value 1-4 columns to the right of the
    label. We scan up to 8 columns out, preferring cells that already
    contain a value of the kind we're about to write (so a date label
    finds the existing-date cell rather than an intervening blank).
    """
    candidates = []
    for c in range(col + 1, min(col + 9, ws.max_column + 2)):
        candidates.append(ws.cell(row=row, column=c))

    # Preferred: an existing value of the matching kind.
    for cell in candidates:
        v = cell.value
        if kind == "date" and isinstance(v, (date, datetime)):
            return cell
        if kind == "invoice_number" and isinstance(v, str) and v.strip():
            # Avoid picking a label-like neighbour (e.g. "Date:").
            if not v.strip().endswith(":"):
                return cell
    # Fallback: first non-empty cell.
    for cell in candidates:
        if cell.value not in (None, ""):
            return cell
    # Last resort: first cell after the label.
    return candidates[0] if candidates else None


def _write_value(cell, kind: str, value) -> None:
    if kind == "date":
        cell.value = value
        # Preserve any existing date format if present; otherwise apply
        # a generic m/d/yyyy.
        if not (cell.number_format and "y" in cell.number_format.lower()):
            cell.number_format = "m/d/yyyy"
    elif kind == "invoice_number":
        cell.value = _format_invoice_number_like(cell.value, value)
    else:
        cell.value = value


def _format_invoice_number_like(existing, new_str: str) -> str:
    """If the existing invoice number used a space (e.g. 'HART 23'), keep
    that style for the new value (e.g. 'HART 24'). Otherwise return the
    new string as-is.
    """
    if not isinstance(existing, str):
        return new_str
    m = re.match(r"^([A-Za-z]+)(\s+)(\d+)\s*$", existing)
    if not m:
        return new_str
    prefix = m.group(1)
    spacing = m.group(2)
    # Extract number portion of new_str; allow letters before digits.
    nm = re.match(r"^([A-Za-z]+)\s*(\d+)\s*$", new_str)
    if nm and nm.group(1).lower() == prefix.lower():
        return f"{prefix}{spacing}{nm.group(2)}"
    return new_str
